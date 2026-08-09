from __future__ import annotations

import json
import base64
import gc
import mimetypes
import shutil
import subprocess
import uuid
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PROJECT = ROOT / "project"
TEMPLATE = PROJECT / "instance_template" / "boatboard.xlsx"
INSTANCE = PROJECT / "private_instance"
WORKBOOK = INSTANCE / "boatboard.xlsx"
BOARD = INSTANCE / "board.json"
IMAGES = INSTANCE / "images"


def ensure_instance() -> None:
    INSTANCE.mkdir(parents=True, exist_ok=True)
    IMAGES.mkdir(parents=True, exist_ok=True)
    if not WORKBOOK.exists():
        shutil.copy2(TEMPLATE, WORKBOOK)
    if not BOARD.exists():
        BOARD.write_text('{"version":1,"teams":{}}\n', encoding="utf-8")


def rows_by_header(sheet, header_row: int = 3) -> list[dict[str, object]]:
    headers = [str(cell.value or "").strip() for cell in sheet[header_row]]
    rows = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def read_organization(path: Path = WORKBOOK) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True)
    company_values = {
        str(row[0].value or "").strip(): row[1].value
        for row in workbook["Company"].iter_rows(min_row=4, max_col=2)
        if row[0].value
    }
    team_rows = rows_by_header(workbook["Teams"])
    colleague_rows = rows_by_header(workbook["Colleagues"])
    teams = [
        {"id": str(row["team_id"]).strip(), "name": str(row["team_name"]).strip()}
        for row in team_rows
        if row.get("team_id") and row.get("team_name")
    ]
    team_ids = {team["id"] for team in teams}
    colleagues = []
    for row in colleague_rows:
        person_id = str(row.get("person_id") or "").strip()
        name = str(row.get("name") or "").strip()
        team_id = str(row.get("team_id") or "").strip()
        if not person_id or not name or team_id not in team_ids:
            continue
        image_filename = Path(str(row.get("image_filename") or "").strip()).name
        colleagues.append({
            "id": person_id,
            "name": name,
            "teamId": team_id,
            "imageUrl": f"/instance-images/{image_filename}" if image_filename else None,
            "role": str(row.get("role") or "").strip(),
            "notes": str(row.get("notes") or "").strip(),
            "whatsapp": str(row.get("whatsapp") or "").strip(),
            "discord": str(row.get("discord") or "").strip(),
            "description": str(row.get("description") or "").strip(),
        })
    organization = {
        "companyName": str(company_values.get("company_name") or "BoatBoard"),
        "pageTitle": str(company_values.get("page_title") or "Boat Board"),
        "teams": teams,
        "colleagues": colleagues,
    }
    workbook.close()
    return organization


def write_organization(payload: dict[str, object]) -> None:
    workbook = load_workbook(WORKBOOK)
    company = workbook["Company"]
    company["B4"] = str(payload.get("companyName") or "BoatBoard")
    company["B5"] = str(payload.get("pageTitle") or "Boat Board")
    teams_sheet = workbook["Teams"]
    colleagues_sheet = workbook["Colleagues"]
    if teams_sheet.max_row >= 4:
        teams_sheet.delete_rows(4, teams_sheet.max_row - 3)
    if colleagues_sheet.max_row >= 4:
        colleagues_sheet.delete_rows(4, colleagues_sheet.max_row - 3)
    for team in payload.get("teams", []):
        teams_sheet.append([team.get("id", ""), team.get("name", ""), ""])
    for person in payload.get("colleagues", []):
        image_filename = Path(str(person.get("imageFilename") or "")).name
        colleagues_sheet.append([
            person.get("id", ""), person.get("name", ""), person.get("teamId", ""),
            image_filename, person.get("role", ""), person.get("notes", ""),
            person.get("whatsapp", ""), person.get("discord", ""), person.get("description", ""),
        ])
    temporary = WORKBOOK.with_name(f"{WORKBOOK.stem}.tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    temporary.replace(WORKBOOK)


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=True)
    required = {
        "Company": {"field", "value"},
        "Teams": {"team_id", "team_name"},
        "Colleagues": {"person_id", "name", "team_id", "image_filename"},
    }
    missing_sheets = set(required) - set(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"Missing sheet(s): {', '.join(sorted(missing_sheets))}")
    for sheet_name, required_headers in required.items():
        headers = {str(cell.value or "").strip() for cell in workbook[sheet_name][3]}
        missing_headers = required_headers - headers
        if missing_headers:
            raise ValueError(f"{sheet_name} is missing column(s): {', '.join(sorted(missing_headers))}")
    workbook.close()
    organization = read_organization(path)
    team_ids = [team["id"] for team in organization["teams"]]
    person_ids = [person["id"] for person in organization["colleagues"]]
    if len(team_ids) != len(set(team_ids)):
        raise ValueError("Team IDs must be unique.")
    if len(person_ids) != len(set(person_ids)):
        raise ValueError("Person IDs must be unique.")


def reconcile_board_file(organization: dict[str, object]) -> None:
    try:
        current = json.loads(BOARD.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        current = {"version": 1, "teams": {}}
    current_teams = current.get("teams", {}) if isinstance(current, dict) else {}
    colleagues = organization["colleagues"]
    colleague_by_id = {person["id"]: person for person in colleagues}
    reconciled = {}
    for team in organization["teams"]:
        stored = current_teams.get(team["id"])
        if not isinstance(stored, dict):
            continue
        valid_team_ids = {
            person["id"] for person in colleagues if person["teamId"] == team["id"]
        }
        stored_order = stored.get("profileOrder", [])
        profile_order = []
        if isinstance(stored_order, list):
            profile_order = list(dict.fromkeys(
                person_id for person_id in stored_order if person_id in valid_team_ids
            ))
        remaining = sorted(
            (person for person in colleagues if person["id"] in valid_team_ids and person["id"] not in profile_order),
            key=lambda person: person["name"].casefold(),
        )
        profile_order.extend(person["id"] for person in remaining)
        leader_id = stored.get("leaderId")
        leader = colleague_by_id.get(leader_id)
        if not leader or leader["teamId"] == team["id"]:
            leader_id = None
        reconciled[team["id"]] = {
            "placed": stored.get("placed") is not False,
            "x": stored.get("x"),
            "y": stored.get("y"),
            "leaderId": leader_id,
            "profileOrder": profile_order,
        }
    temporary = BOARD.with_suffix(".json.tmp")
    temporary.write_text(json.dumps({"version": 1, "teams": reconciled}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(BOARD)


class BoatBoardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PROJECT), **kwargs)

    def send_json(self, payload: object, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        try:
            if path == "/api/organization":
                self.send_json(read_organization())
                return
            if path == "/api/board":
                self.send_json(json.loads(BOARD.read_text(encoding="utf-8-sig")))
                return
            if path.startswith("/instance-images/"):
                filename = Path(path.removeprefix("/instance-images/")).name
                image_path = IMAGES / filename
                if not image_path.is_file():
                    self.send_error(404)
                    return
                content = image_path.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", mimetypes.guess_type(image_path.name)[0] or "application/octet-stream")
                self.send_header("Content-Length", str(len(content)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(content)
                return
            super().do_GET()
        except Exception as error:
            self.send_json({"error": str(error)}, 500)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        if path not in ("/api/board", "/api/organization", "/api/image", "/api/workbook", "/api/open-instance-folder"):
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if path == "/api/open-instance-folder":
                subprocess.Popen(["explorer.exe", str(INSTANCE)])
                self.send_json({"opened": True})
                return
            if path == "/api/workbook":
                filename = str(payload.get("filename") or "")
                if not filename.lower().endswith(".xlsx"):
                    raise ValueError("Choose an .xlsx workbook.")
                candidate = INSTANCE / f"boatboard.candidate-{uuid.uuid4().hex}.xlsx"
                replacement = INSTANCE / "boatboard.replacement.xlsx"
                workbook_bytes = base64.b64decode(payload.get("data") or "")
                candidate.write_bytes(workbook_bytes)
                try:
                    validate_workbook(candidate)
                    gc.collect()
                    backup = INSTANCE / "boatboard.previous.xlsx"
                    if WORKBOOK.exists():
                        shutil.copy2(WORKBOOK, backup)
                    replacement.write_bytes(workbook_bytes)
                    replacement.replace(WORKBOOK)
                    reconcile_board_file(read_organization())
                finally:
                    try:
                        candidate.unlink(missing_ok=True)
                    except PermissionError:
                        pass
                    replacement.unlink(missing_ok=True)
                self.send_json({"saved": True})
                return
            if path == "/api/organization":
                write_organization(payload)
                reconcile_board_file(read_organization())
                self.send_json({"saved": True})
                return
            if path == "/api/image":
                filename = Path(str(payload.get("filename") or "")).name
                if not filename:
                    raise ValueError("Image filename is required.")
                (IMAGES / filename).write_bytes(base64.b64decode(payload.get("data") or ""))
                self.send_json({"saved": True, "filename": filename})
                return
            temporary = BOARD.with_suffix(".json.tmp")
            temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            temporary.replace(BOARD)
            self.send_json({"saved": True})
        except Exception as error:
            self.send_json({"error": str(error)}, 400)


if __name__ == "__main__":
    ensure_instance()
    server = ThreadingHTTPServer(("127.0.0.1", 4173), BoatBoardHandler)
    print("BoatBoard: http://127.0.0.1:4173/")
    print("Editor:    http://127.0.0.1:4173/editor.html")
    server.serve_forever()
